import os
import shutil
import zipfile
from datetime import datetime

import openpyxl
from sqlalchemy.orm import Session

from app.database import SQLALCHEMY_DATABASE_URL, SessionLocal
from app.models import Pattuglia, Prenotazione

BACKUP_DIR = os.getenv("BACKUP_DIR", "data/backups")
GOOGLE_DRIVE_FOLDER_ID = os.getenv("GOOGLE_DRIVE_FOLDER_ID", "")

# Ensure backup dir exists
os.makedirs(BACKUP_DIR, exist_ok=True)


def generate_excel_sfide(db: Session, filepath: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    if not ws:
        ws = wb.create_sheet()
    ws.title = "Sfide e Punteggi"

    # Header
    ws.append(["Unita", "Pattuglia", "Punteggio Attuale", "Capo Pattuglia"])

    pattuglie = db.query(Pattuglia).all()
    for p in pattuglie:
        ws.append([p.unita.name if p.unita else "N/A", p.name, p.current_score, p.capo_pattuglia])

    wb.save(filepath)


def generate_excel_riservazioni(db: Session, filepath: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    if not ws:
        ws = wb.create_sheet()
    ws.title = "Riservazioni Terreni"

    # Header
    ws.append(["Terreno", "Unita", "Inizio", "Fine", "Durata (h)", "Stato", "Note"])

    prenotazioni = db.query(Prenotazione).all()
    for p in prenotazioni:
        terreno_name = p.terreno.name if p.terreno else "N/A"
        unita_name = p.unita.name if p.unita else "N/A"
        start = p.start_time.strftime("%Y-%m-%d %H:%M") if p.start_time else ""
        end = p.end_time.strftime("%Y-%m-%d %H:%M") if p.end_time else ""

        ws.append([terreno_name, unita_name, start, end, p.duration, p.status, p.notes or ""])

    wb.save(filepath)


def cleanup_old_backups(backup_dir: str, max_backups: int = 10):
    """Keep only the most recent max_backups backup runs."""
    try:
        if not os.path.exists(backup_dir):
            return
        files = os.listdir(backup_dir)
        runs = {}
        for f in files:
            parts = f.split("_")
            if len(parts) >= 2:
                prefix = f"{parts[0]}_{parts[1]}"
                if len(parts[0]) == 8 and parts[0].isdigit() and len(parts[1]) == 4 and parts[1].isdigit():
                    if prefix not in runs:
                        runs[prefix] = []
                    runs[prefix].append(f)

        sorted_prefixes = sorted(runs.keys())
        if len(sorted_prefixes) > max_backups:
            prefixes_to_delete = sorted_prefixes[:-max_backups]
            for prefix in prefixes_to_delete:
                for f in runs[prefix]:
                    file_path = os.path.join(backup_dir, f)
                    try:
                        os.remove(file_path)
                    except Exception as e:
                        print(f"Failed to delete old backup file {file_path}: {e}")
    except Exception as e:
        print(f"Cleanup of old backups failed: {e}")


def execute_backup() -> tuple[bool, str, str | None]:
    """
    Executes the full backup process:
    1. Copies DB
    2. Generates Excels
    3. Zips database and Excels into a single ZIP archive
    4. Cleans up old backups keeping a rolling window of 10 backups
    Returns (Success, message, zip_path)
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    db_path = SQLALCHEMY_DATABASE_URL.replace("sqlite:///", "")
    zip_path = None

    try:
        # 1. DB Copy
        backup_db_path = os.path.join(BACKUP_DIR, f"{timestamp}_backup_db.sqlite")
        if os.path.exists(db_path):  # pragma: no cover
            shutil.copy2(db_path, backup_db_path)

        # 2. Excels
        db = SessionLocal()
        try:
            sfide_path = os.path.join(BACKUP_DIR, f"{timestamp}_sfide_e_punteggi.xlsx")
            generate_excel_sfide(db, sfide_path)

            riserv_path = os.path.join(BACKUP_DIR, f"{timestamp}_riservazioni_terreni.xlsx")
            generate_excel_riservazioni(db, riserv_path)
        finally:
            db.close()

        # 3. Create ZIP archive
        zip_path = os.path.join(BACKUP_DIR, f"{timestamp}_backup.zip")
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zip_file:
            if os.path.exists(backup_db_path):
                zip_file.write(backup_db_path, arcname="punteggiometro.sqlite")
            if os.path.exists(sfide_path):
                zip_file.write(sfide_path, arcname="sfide_e_punteggi.xlsx")
            if os.path.exists(riserv_path):
                zip_file.write(riserv_path, arcname="riservazioni_terreni.xlsx")

        # 4. Rolling window cleanup
        cleanup_old_backups(BACKUP_DIR, max_backups=10)

        return True, "Backup completato con successo", zip_path
    except Exception as e:
        print(f"Backup failed: {e}")
        return False, f"Errore durante il backup: {str(e)}", None
